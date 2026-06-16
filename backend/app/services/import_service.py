from __future__ import annotations

import re
import warnings
from datetime import date, datetime, time
from io import BytesIO
from typing import Any

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from ..auth import CurrentUser
from ..schemas import ActivationCreate, ImportResult
from .activation_service import create_activation

EXPECTED_COLUMNS = {
    "modelo da moto": "motorcycle_model",
    "chassi": "chassis",
    "data de pedido": "order_date",
    "vendedor": "seller_responsible_name",
    "data de ativacao": "activation_date",
    "horario": "activation_time",
    "nome do cliente": "client_name",
    "cpf do cliente": "client_cpf",
    "observacoes": "notes",
}

CPF_PATTERN = re.compile(
    r"(\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2}|\d{3}\.?\d{3}\.?\d{3}-?\d{2}|\d{11,14})"
)
DATE_PATTERN = re.compile(r"(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?")


def normalize_header(header: Any) -> str:
    return str(header or "").strip().lower()


def parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and 30000 <= float(value) <= 60000:
        return from_excel(value).date()
    value_str = str(value).strip()
    for date_format in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(value_str, date_format).date()
        except ValueError:
            continue
    raise ValueError(f"Data inválida: {value_str}")


def safe_parse_date(value: Any) -> date | None:
    try:
        return parse_date(value)
    except Exception:
        return None


def parse_time(value: Any) -> time:
    if value in (None, ""):
        raise ValueError("Horário vazio.")
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    value_str = str(value).strip()
    for time_format in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(value_str, time_format).time().replace(second=0, microsecond=0)
        except ValueError:
            continue
    raise ValueError(f"Horário inválido: {value_str}")


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def is_date_only_row(row: list[Any]) -> bool:
    if not row:
        return False
    first_value = row[0]
    if not isinstance(first_value, (datetime, date)):
        return False
    return all(is_blank(value) for value in row[1:8])


def is_header_row(row: list[Any]) -> bool:
    normalized = [normalize_header(value) for value in row[:8] if not is_blank(value)]
    joined = " | ".join(normalized)
    return "modelo" in joined and "chassi" in joined and "vendedor" in joined


def parse_time_fragment(value: Any) -> time | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        parsed_time = value.time().replace(second=0, microsecond=0)
        if parsed_time != time(0, 0):
            return parsed_time
        return None
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)

    text = str(value).strip().upper()
    if not text:
        return None
    text = text.replace("O", "0")
    text = text.replace("HORAS", "H").replace("HRS", "H").replace("HR", "H")
    text = text.replace("ÀS", " ").replace("AS", " ").replace("---", " ").replace("-", " ")
    text = " ".join(text.split())

    match = re.search(r"(\d{1,2})\s*:\s*(\d{2})", text)
    if match:
        hours = int(match.group(1))
        minutes = int(match.group(2))
        return time(hours % 24, minutes % 60)

    match = re.search(r"(\d{1,2})\s*H\s*(\d{2})", text)
    if match:
        hours = int(match.group(1))
        minutes = int(match.group(2))
        return time(hours % 24, minutes % 60)

    match = re.search(r"(\d{1,2})\s*H\b", text)
    if match:
        hours = int(match.group(1))
        return time(hours % 24, 0)

    match = re.search(r"\b(\d{1,2})(\d{2})\b", text)
    if match and len(text) <= 6:
        hours = int(match.group(1))
        minutes = int(match.group(2))
        return time(hours % 24, minutes % 60)

    return None


def parse_schedule(value: Any, reference_date: date | None = None) -> tuple[date | None, time | None]:
    if value in (None, ""):
        return reference_date, None
    if isinstance(value, datetime):
        return value.date(), value.time().replace(second=0, microsecond=0)
    if isinstance(value, date):
        return value, None

    text = str(value).strip()
    if not text:
        return reference_date, None

    match = DATE_PATTERN.search(text)
    activation_date = reference_date
    remainder = text
    if match:
        day = int(match.group(1))
        month = int(match.group(2))
        year_text = match.group(3)
        if year_text:
            year = int(year_text)
            if year < 100:
                year += 2000
        elif reference_date:
            year = reference_date.year
        else:
            year = datetime.now().year
        try:
            activation_date = date(year, month, day)
        except ValueError:
            activation_date = reference_date
        remainder = (text[:match.start()] + " " + text[match.end():]).strip()

    activation_time = parse_time_fragment(remainder)
    return activation_date, activation_time


def parse_client_and_cpf(value: Any) -> tuple[str, str]:
    text = str(value or "").replace("\n", " ").strip()
    text = " ".join(text.split())
    if not text:
        return "Cliente não informado", "NAO INFORMADO"

    match = CPF_PATTERN.search(text)
    if not match:
        digits = "".join(char for char in text if char.isdigit())
        if len(digits) >= 11:
            return "Cliente não informado", digits[:14]
        return text.strip(" -") or "Cliente não informado", "NAO INFORMADO"

    cpf = match.group(1)
    name = f"{text[:match.start()]} {text[match.end():]}".strip()
    name = re.sub(r"(?i)\bCPF\b", " ", name)
    name = re.sub(r"\s*-\s*", " ", name)
    name = " ".join(name.split()).strip(" -")
    if not name:
        name = "Cliente não informado"
    return name, cpf


def normalize_seller_name(value: Any) -> str:
    seller_name = str(value or "").strip()
    return seller_name or "Não informado"


def parse_legacy_row(
    row: list[Any],
    *,
    section_date: date | None,
) -> ActivationCreate:
    order_date = safe_parse_date(row[2])
    activation_date, activation_time = parse_schedule(row[4], reference_date=section_date or order_date)
    client_name, client_cpf = parse_client_and_cpf(row[5])

    notes_parts: list[str] = []
    observation = str(row[6] or "").strip() if len(row) > 6 else ""
    if observation:
        notes_parts.append(observation)
    if activation_time is None:
        activation_time = time(0, 0)
        notes_parts.append("Horário original não identificado na planilha legada.")
    if activation_date is None:
        raise ValueError("Não foi possível identificar a data de ativação.")

    return ActivationCreate(
        motorcycle_model=str(row[0] or "").strip(),
        chassis=str(row[1] or "").strip(),
        order_date=order_date,
        seller_responsible_name=normalize_seller_name(row[3]),
        activation_date=activation_date,
        activation_time=activation_time,
        client_name=client_name,
        client_cpf=client_cpf,
        notes=" | ".join(part for part in notes_parts if part),
        mechanic_notes="",
        status="Pendente",
    )


def parse_modern_row(
    row: list[Any],
    *,
    section_date: date | None,
) -> ActivationCreate:
    order_date = safe_parse_date(row[2])
    activation_date = safe_parse_date(row[4]) or section_date or order_date
    activation_time = parse_time_fragment(row[5])
    if activation_time is None and isinstance(row[4], datetime):
        activation_time = row[4].time().replace(second=0, microsecond=0)
        if activation_time == time(0, 0):
            activation_time = None

    notes_parts: list[str] = []
    if activation_time is None:
        activation_time = time(0, 0)
        notes_parts.append("Horário original não identificado na planilha legada.")

    return ActivationCreate(
        motorcycle_model=str(row[0] or "").strip(),
        chassis=str(row[1] or "").strip(),
        order_date=order_date,
        seller_responsible_name=normalize_seller_name(row[3]),
        activation_date=activation_date,
        activation_time=activation_time,
        client_name=str(row[6] or "").strip() or "Cliente não informado",
        client_cpf=str(row[7] or "").strip() or "NAO INFORMADO",
        notes=" | ".join(notes_parts),
        mechanic_notes="",
        status="Pendente",
    )


def parse_hybrid_row(
    row: list[Any],
    *,
    section_date: date | None,
) -> ActivationCreate:
    order_date = safe_parse_date(row[2])
    activation_date = safe_parse_date(row[4]) or section_date or order_date
    activation_time = parse_time_fragment(row[5]) or time(0, 0)
    client_name, client_cpf = parse_client_and_cpf(row[6])

    notes_parts: list[str] = []
    if parse_time_fragment(row[5]) is None:
        notes_parts.append("Horário original não identificado na planilha legada.")

    return ActivationCreate(
        motorcycle_model=str(row[0] or "").strip(),
        chassis=str(row[1] or "").strip(),
        order_date=order_date,
        seller_responsible_name=normalize_seller_name(row[3]),
        activation_date=activation_date,
        activation_time=activation_time,
        client_name=client_name,
        client_cpf=client_cpf,
        notes=" | ".join(notes_parts),
        mechanic_notes="",
        status="Pendente",
    )


def parse_workbook_rows(workbook: Any) -> tuple[list[tuple[str, int, ActivationCreate]], int, list[str]]:
    parsed_rows: list[tuple[str, int, ActivationCreate]] = []
    blank_rows = 0
    parse_errors: list[str] = []

    for worksheet in workbook.worksheets:
        section_date: date | None = None
        for row_number, raw_row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            row = list(raw_row[:8])
            if all(is_blank(value) for value in row):
                blank_rows += 1
                continue
            if is_header_row(row):
                continue
            if is_date_only_row(row):
                section_date = parse_date(row[0])
                continue
            if is_blank(row[0]):
                continue

            has_separate_client_fields = len(row) >= 8 and not is_blank(row[6]) and not is_blank(row[7])
            has_hybrid_fields = (
                not has_separate_client_fields
                and len(row) >= 7
                and not is_blank(row[6])
                and isinstance(row[4], (datetime, date))
                and parse_time_fragment(row[5]) is not None
            )
            try:
                if has_separate_client_fields:
                    payload = parse_modern_row(row, section_date=section_date)
                elif has_hybrid_fields:
                    payload = parse_hybrid_row(row, section_date=section_date)
                else:
                    payload = parse_legacy_row(row, section_date=section_date)
                parsed_rows.append((worksheet.title, row_number, payload))
            except Exception as exc:
                parse_errors.append(f"Aba {worksheet.title}, linha {row_number}: {exc}")

    return parsed_rows, blank_rows, parse_errors


def import_excel(file_bytes: bytes, current_user: CurrentUser) -> ImportResult:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)

    parsed_rows, ignored_blank_rows, errors = parse_workbook_rows(workbook)
    if not parsed_rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Não foi possível localizar linhas válidas na planilha.")

    imported = 0

    for sheet_name, row_number, payload in parsed_rows:
        try:
            create_activation(payload, current_user)
            imported += 1
        except Exception as exc:
            errors.append(f"Aba {sheet_name}, linha {row_number}: {exc}")

    return ImportResult(imported=imported, ignored_blank_rows=ignored_blank_rows, errors=errors)


async def import_upload_file(upload_file: UploadFile, current_user: CurrentUser) -> ImportResult:
    if not upload_file.filename:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Informe um arquivo para importação.")
    if not upload_file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Somente arquivos .xlsx são suportados.")

    file_bytes = await upload_file.read()
    return import_excel(file_bytes, current_user)
